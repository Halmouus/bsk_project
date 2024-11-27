from django.urls import reverse_lazy
from django.db import models
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from .models import Invoice, InvoiceProduct, Product, ExportRecord, Check, Supplier
from .forms import InvoiceCreateForm, InvoiceUpdateForm  # Import the custom form here
from django.forms import inlineformset_factory
from django.contrib.messages.views import SuccessMessageMixin
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
from django.contrib.auth.mixins import UserPassesTestMixin
from django.db.models import Q,F, Case, When, DecimalField, Subquery, Sum, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.contrib import messages
from decimal import Decimal, InvalidOperation
from django.template.loader import render_to_string
from django.db.models.sql.where import EmptyResultSet



@method_decorator(csrf_exempt, name='dispatch')
class AddProductToInvoiceView(View):
    def post(self, request):
        invoice_id = request.POST.get('invoice_id')
        product_id = request.POST.get('product')
        quantity = request.POST.get('quantity')
        unit_price = request.POST.get('unit_price')
        vat_rate = request.POST.get('vat_rate')
        reduction_rate = request.POST.get('reduction_rate', 0)  # Add default value
        expense_code = request.POST.get('expense_code')

        try:
            # Fetch the invoice and product
            invoice = get_object_or_404(Invoice, pk=invoice_id)
            product = get_object_or_404(Product, pk=product_id)

            # Create a new InvoiceProduct entry
            invoice_product = InvoiceProduct.objects.create(
                invoice=invoice,
                product=product,
                quantity=quantity,
                unit_price=unit_price,
                vat_rate=vat_rate,
                reduction_rate=reduction_rate
            )

            # Success response
            return JsonResponse({"message": "Product added successfully."}, status=200)
        
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

# List all Invoices
class InvoiceListView(ListView):
    model = Invoice
    template_name = 'invoice/invoice_list.html'
    context_object_name = 'invoices'

    def get_queryset(self):
        queryset = Invoice.objects.all().select_related('supplier').prefetch_related('products')

        # Debug prints
        print("Request GET params:", self.request.GET)

        # Date Range Filter
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        try:
            if date_from:
                queryset = queryset.filter(date__gte=date_from)
            if date_to:
                queryset = queryset.filter(date__lte=date_to)
        except Exception as e:
            print(f"Error filtering by date range: {e}")

        # Amount Range Filter
        amount_min = self.request.GET.get('amount_min')
        amount_max = self.request.GET.get('amount_max')
        try:
            if amount_min or amount_max:
                invoices = list(queryset)  # Evaluate queryset into a list for manual filtering
                filtered_invoices = []

                amount_min = Decimal(amount_min if amount_min else '0')
                amount_max = Decimal(amount_max if amount_max else '999999999')

                for invoice in invoices:
                    net_amount = invoice.net_amount  # Assume net_amount is a computed property
                    if amount_min <= net_amount <= amount_max:
                        filtered_invoices.append(invoice.id)

                queryset = queryset.filter(id__in=filtered_invoices)
        except Exception as e:
            print(f"Error filtering by amount range: {e}")

        # Supplier Filter
        supplier = self.request.GET.get('supplier')
        try:
            if supplier:
                queryset = queryset.filter(supplier_id=supplier)
        except Exception as e:
            print(f"Error filtering by supplier: {e}")

        # Payment Status Filter
        payment_status = self.request.GET.get('payment_status')
        try:
            if payment_status:
                queryset = queryset.filter(payment_status=payment_status)
        except Exception as e:
            print(f"Error filtering by payment status: {e}")

        # Export Status Filter
        export_status = self.request.GET.get('export_status')
        try:
            if export_status == 'exported':
                queryset = queryset.filter(exported_at__isnull=False)
            elif export_status == 'not_exported':
                queryset = queryset.filter(exported_at__isnull=True)
        except Exception as e:
            print(f"Error filtering by export status: {e}")

        # Product Filter
        product_id = self.request.GET.get('product')
        try:
            if product_id:
                queryset = queryset.filter(products__product_id=product_id)
        except Exception as e:
            print(f"Error filtering by product: {e}")

        # Payment Status Filters
        try:
            has_pending_checks = self.request.GET.get('has_pending_checks')
            if has_pending_checks:
                queryset = queryset.filter(check__status='pending').distinct()

            has_delivered_unpaid = self.request.GET.get('has_delivered_unpaid')
            if has_delivered_unpaid:
                queryset = queryset.filter(check__status='delivered').exclude(check__status='paid').distinct()
        except Exception as e:
            print(f"Error filtering by payment status checks: {e}")

        # Energy Filter
        is_energy = self.request.GET.get('is_energy')
        try:
            if is_energy:
                queryset = queryset.filter(supplier__is_energy=True)
        except Exception as e:
            print(f"Error filtering by energy suppliers: {e}")

        # Credit Note Status
        credit_note_status = self.request.GET.get('credit_note_status')
        try:
            if credit_note_status == 'has_credit_notes':
                queryset = queryset.filter(credit_notes__isnull=False).distinct()
            elif credit_note_status == 'no_credit_notes':
                queryset = queryset.filter(credit_notes__isnull=True)
            elif credit_note_status == 'partially_credited':
                queryset = queryset.filter(
                    credit_notes__isnull=False,
                    payment_status__in=['not_paid', 'partially_paid']
                ).distinct()
        except Exception as e:
            print(f"Error filtering by credit note status: {e}")

        # Due Date Range
        due_date_from = self.request.GET.get('due_date_from')
        due_date_to = self.request.GET.get('due_date_to')
        try:
            if due_date_from:
                queryset = queryset.filter(payment_due_date__gte=due_date_from)
            if due_date_to:
                queryset = queryset.filter(payment_due_date__lte=due_date_to)
        except Exception as e:
            print(f"Error filtering by due date range: {e}")

        # Overdue Filter
        is_overdue = self.request.GET.get('is_overdue')
        try:
            if is_overdue:
                today = timezone.now().date()
                queryset = queryset.filter(
                    payment_due_date__lt=today,
                    payment_status__in=['not_paid', 'partially_paid']
                )
        except Exception as e:
            print(f"Error filtering by overdue invoices: {e}")

        # Print final queryset SQL for debugging
        try:
            print("Final query SQL:", queryset.query)
        except Exception as e:
            print(f"Error printing final query SQL: {e}")

        return queryset.order_by('-date')


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add filter counts
        active_filters = {}
        
        # Date Range
        if self.request.GET.get('date_from') or self.request.GET.get('date_to'):
            date_range = []
            if self.request.GET.get('date_from'):
                date_range.append(f"From: {self.request.GET.get('date_from')}")
            if self.request.GET.get('date_to'):
                date_range.append(f"To: {self.request.GET.get('date_to')}")
            active_filters['date_range'] = ' - '.join(date_range)

        # Supplier
        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            try:
                supplier = Supplier.objects.get(id=supplier_id)
                active_filters['supplier'] = supplier.name
            except Supplier.DoesNotExist:
                pass

        # Payment Status
        payment_status = self.request.GET.get('payment_status')
        if payment_status:
            status_display = {
                'not_paid': 'Not Paid',
                'partially_paid': 'Partially Paid',
                'paid': 'Paid'
            }
            active_filters['payment_status'] = status_display.get(payment_status)

        # Amount Range
        if self.request.GET.get('amount_min') or self.request.GET.get('amount_max'):
            amount_range = []
            if self.request.GET.get('amount_min'):
                amount_range.append(f"Min: {self.request.GET.get('amount_min')}")
            if self.request.GET.get('amount_max'):
                amount_range.append(f"Max: {self.request.GET.get('amount_max')}")
            active_filters['amount_range'] = ' - '.join(amount_range)

        # Export Status
        export_status = self.request.GET.get('export_status')
        if export_status:
            active_filters['export_status'] = 'Exported' if export_status == 'exported' else 'Not Exported'

        # Document Type
        document_type = self.request.GET.get('document_type')
        if document_type:
            active_filters['document_type'] = 'Invoice' if document_type == 'invoice' else 'Credit Note'

        context['active_filters'] = active_filters
        context['total_results'] = self.get_queryset().count()
        
        # Add initial supplier data for the filter if selected
        if supplier_id:
            try:
                supplier = Supplier.objects.get(id=supplier_id)
                context['initial_supplier'] = {
                    'id': supplier_id,
                    'text': supplier.name
                }
            except Supplier.DoesNotExist:
                pass

        return context

    def render_to_response(self, context, **response_kwargs):
        """Handle both HTML and AJAX responses"""
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'html': render_to_string(
                    'invoice/partials/invoice_table.html',
                    context,
                    request=self.request
                ),
                'total_results': context['total_results'],
                'active_filters': context['active_filters']
            })
        return super().render_to_response(context, **response_kwargs)

# Create a new Invoice
class InvoiceCreateView(SuccessMessageMixin, CreateView):
    model = Invoice
    form_class = InvoiceUpdateForm  # Use the custom form here
    template_name = 'invoice/invoice_form.html'
    success_url = reverse_lazy('invoice-list')
    success_message = "Invoice successfully created."

    def form_valid(self, form):
        response = super().form_valid(form)
        # We may want to pass the newly created invoice to the next page or modal
        return response

    def get_form_class(self):
        print("Using CREATE VIEW")  # Debug print
        return InvoiceCreateForm

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        data['products'] = Product.objects.all()  # Add all products to the context for dropdown population
        return data

# Update an existing Invoice
class InvoiceUpdateView(SuccessMessageMixin, UpdateView):
    model = Invoice
    form_class = InvoiceUpdateForm
    template_name = 'invoice/invoice_form.html'
    success_url = reverse_lazy('invoice-list')
    success_message = "Invoice successfully updated."

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs) 
        if self.request.POST:
            data['products'] = InvoiceProductInlineFormset(self.request.POST, instance=self.object) 
        else:
             data['products'] = InvoiceProductInlineFormset(instance=self.object, queryset=InvoiceProduct.objects.filter(invoice=self.object))
        return data


    def get_form_class(self):
        print("Using UPDATE VIEW")  # Debug print
        return InvoiceUpdateForm
    
    def form_valid(self, form):
        print("Entering form_valid")
        print("Form data:", form.cleaned_data)
        context = self.get_context_data()
        products = context['products']
        print("Form valid:", form.is_valid())
        print("Products valid:", products.is_valid())
        if not products.is_valid():
            print("Products errors:", products.errors)  # Add this
            print("Non-form errors:", products.non_form_errors())  # And this
        if form.is_valid() and products.is_valid():
            print("Both form and products are valid")
            self.object = form.save()
            products.instance = self.object
            products.save()
            print("Save completed")
            return super().form_valid(form)
        print("Form validation failed")
        return self.form_invalid(form)

    def dispatch(self, request, *args, **kwargs):
        invoice = self.get_object()
        if invoice.payment_status == 'paid':
            messages.error(request, '<i class="fas fa-lock"></i> This invoice has been paid and cannot be edited!', 
                         extra_tags='danger')
            return redirect('invoice-list')
        if invoice.exported_at:
            messages.error(request, '<i class="fas fa-lock"></i> This invoice has been exported and cannot be edited!', 
                         extra_tags='danger')
            return redirect('invoice-list')
        return super().dispatch(request, *args, **kwargs)

# Delete an Invoice
class InvoiceDeleteView(DeleteView):
    model = Invoice
    template_name = 'invoice/invoice_confirm_delete.html'
    success_url = reverse_lazy('invoice-list')
    success_message = "Invoice successfully deleted."

    def dispatch(self, request, *args, **kwargs):
        invoice = self.get_object()
        if invoice.exported_at:
            messages.error(request, '<i class="fas fa-lock"></i> This invoice has been exported and cannot be deleted!', extra_tags='danger')
            return redirect('invoice-list')
        return super().dispatch(request, *args, **kwargs)


InvoiceProductInlineFormset = inlineformset_factory(
    Invoice, InvoiceProduct,
    fields=['product', 'quantity', 'unit_price', 'reduction_rate', 'vat_rate'],
    extra=1,  # Number of empty forms to display
    can_delete=True
)

# Invoice details view for AJAX request
class InvoiceDetailsView(View):
    def get(self, request):
        invoice_id = request.GET.get('invoice_id')
        try:
            invoice = Invoice.objects.get(pk=invoice_id)
            products = invoice.products.all()
            product_data = [
                {
                    'name': product.product.name,
                    'unit_price': f"{product.unit_price:,.2f}",
                    'quantity': product.quantity,
                    'vat_rate': f"{product.vat_rate}%",  # Add VAT Rate
                    'reduction_rate': product.reduction_rate,
                    'raw_price': f"{product.quantity * product.unit_price * (1 - product.reduction_rate / 100):,.2f}",
                } for product in products
            ]

            # Calculate total raw amount
            total_raw_amount = sum([
                product.quantity * product.unit_price * (1 - product.reduction_rate / 100)
                for product in products
            ])

            # Calculate subtotal per VAT rate
            vat_subtotals = {}
            for product in products:
                vat_rate = product.vat_rate
                raw_price = product.quantity * product.unit_price * (1 - product.reduction_rate / 100)
                if vat_rate not in vat_subtotals:
                    vat_subtotals[vat_rate] = 0
                vat_subtotals[vat_rate] += raw_price * (vat_rate / 100)

            response_data = {
                'products': product_data,
                'total_raw_amount': f"{total_raw_amount:,.2f}",  # Add Total Raw Amount
                'vat_subtotals': [{'vat_rate': f"{rate}%", 'subtotal': f"{subtotal:,.2f}"} for rate, subtotal in vat_subtotals.items()],  # Add VAT Subtotals
                'total_vat': f"{invoice.total_tax_amount:,.2f}",
                'total_amount': f"{invoice.total_amount:,.2f}",
            }
            return JsonResponse(response_data)
        except Invoice.DoesNotExist:
            return JsonResponse({'error': 'Invoice not found'}, status=404)

# Product Autocomplete View
def product_autocomplete(request):
    query = request.GET.get('term', '')
    products = Product.objects.filter(
        Q(name__icontains=query) | 
        Q(fiscal_label__icontains=query)
    )[:10]
    
    product_list = [{
        "label": f"{product.name} ({product.fiscal_label})",
        "value": product.id
    } for product in products]
    
    if not products:
        product_list.append({
            "label": f"Create new product: {query}",
            "value": "new"
        })
        
    return JsonResponse(product_list, safe=False)

@method_decorator(csrf_exempt, name='dispatch')
class EditProductInInvoiceView(View):
    def get(self, request, pk):
        """
        Handles loading the product data for editing.
        """
        try:
            # Fetch the existing InvoiceProduct
            invoice_product = get_object_or_404(InvoiceProduct, pk=pk)
            # Prepare product data to return
            product_data = {
                'product': invoice_product.product.name,
                'product_name': invoice_product.product.name,
                'quantity': invoice_product.quantity,
                'unit_price': float(invoice_product.unit_price),
                'vat_rate': float(invoice_product.vat_rate),
                'reduction_rate': float(invoice_product.reduction_rate),
                'expense_code': invoice_product.product.expense_code,
                'fiscal_label': invoice_product.product.fiscal_label 
            }
            return JsonResponse(product_data, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    def post(self, request, pk):
        """
        Handles updating the product information.
        """
        quantity = request.POST.get('quantity')
        unit_price = request.POST.get('unit_price')
        vat_rate = request.POST.get('vat_rate')
        reduction_rate = request.POST.get('reduction_rate')

        try:
            # Fetch the existing InvoiceProduct
            invoice_product = get_object_or_404(InvoiceProduct, pk=pk)

            # Update the fields with the provided data
            invoice_product.quantity = quantity
            invoice_product.unit_price = unit_price
            invoice_product.vat_rate = vat_rate
            invoice_product.reduction_rate = reduction_rate
            invoice_product.save()

            # Success response
            return JsonResponse({"message": "Product updated successfully."}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    def delete(self, request, pk):
        """
        Handles deleting the product from the invoice.
        """
        try:
            # Fetch the InvoiceProduct instance
            invoice_product = get_object_or_404(InvoiceProduct, pk=pk)

            # Delete the instance
            invoice_product.delete()

            # Success response
            return JsonResponse({"message": "Product deleted successfully."}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@method_decorator(csrf_exempt, name='dispatch')
class ExportInvoicesView(UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.has_perm('testapp.can_export_invoice')

    def generate_excel(self, invoices):
        wb = Workbook()
        ws = wb.active
        ws.title = "Accounting Entries"

        # Define styles
        header_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='344960', end_color='344960', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

        # Set headers
        headers = ['Date', 'Label', 'Debit', 'Credit', 'Account Code', 'Reference', 'Journal', 'Counterpart']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = header_style['border']

        # Set column widths
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 40  # Label
        ws.column_dimensions['C'].width = 15  # Debit
        ws.column_dimensions['D'].width = 15  # Credit
        ws.column_dimensions['E'].width = 15  # Account Code
        ws.column_dimensions['F'].width = 15  # Reference
        ws.column_dimensions['G'].width = 10  # Journal
        ws.column_dimensions['H'].width = 15  # Counterpart

        current_row = 2
        for invoice in invoices:
            entries = invoice.get_accounting_entries()
            for entry in entries:
                ws.cell(row=current_row, column=1, value=entry['date'].strftime('%d/%m/%Y'))
                ws.cell(row=current_row, column=2, value=entry['label'])
                ws.cell(row=current_row, column=3, value=entry['debit'])
                ws.cell(row=current_row, column=4, value=entry['credit'])
                ws.cell(row=current_row, column=5, value=entry['account_code'])
                ws.cell(row=current_row, column=6, value=entry['reference'])
                ws.cell(row=current_row, column=7, value=entry['journal'])
                ws.cell(row=current_row, column=8, value=entry['counterpart'])

                # Style number cells
                for col in [3, 4]:  # Debit and Credit columns
                    cell = ws.cell(row=current_row, column=col)
                    cell.number_format = '# ##0.00'
                    cell.alignment = Alignment(horizontal='right')

                current_row += 1

        return wb

    def post(self, request):
        try:
            data = json.loads(request.body)
            invoice_ids = data.get('invoice_ids', [])
            invoices = Invoice.objects.filter(id__in=invoice_ids, exported_at__isnull=True)

            if not invoices:
                return JsonResponse({'error': 'No valid invoices to export'}, status=400)

            # Generate Excel file
            wb = self.generate_excel(invoices)

            # Create export record
            export_record = ExportRecord.objects.create(
                exported_by=request.user,
                filename=f'accounting_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )

            # Mark invoices as exported
            for invoice in invoices:
                invoice.exported_at = timezone.now()
                invoice.export_history.add(export_record)
                invoice.save()

            # Prepare response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{export_record.filename}"'
            wb.save(response)

            return response

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@method_decorator(csrf_exempt, name='dispatch')
class UnexportInvoiceView(UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.has_perm('testapp.can_unexport_invoice')

    def post(self, request, invoice_id):
        try:
            invoice = get_object_or_404(Invoice, id=invoice_id)
            if not invoice.exported_at:
                return JsonResponse({'error': 'Invoice is not exported'}, status=400)

            # Create export record for the unexport action
            ExportRecord.objects.create(
                exported_by=request.user,
                filename=f'unexport_{invoice.ref}_{timezone.now().strftime("%Y%m%d_%H%M%S")}',
                note=f'Unexported by {request.user.username}'
            )

            # Clear export date
            invoice.exported_at = None
            invoice.save()

            return JsonResponse({'message': 'Invoice successfully unexported'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
        
class InvoicePaymentDetailsView(View):
    def get(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        payment_details = invoice.get_payment_details()
        
        # Get all related checks with their details
        checks = Check.objects.filter(cause=invoice).select_related('checker')
        check_details = [{
            'id': str(check.id),
            'reference': f"{getattr(check.checker.bank_account, 'bank', 'Unknown')}-{check.position}",
            'amount': float(check.amount),
            'status': check.status,
            'created_at': check.creation_date.strftime('%Y-%m-%d'),
            'delivered_at': check.delivered_at.strftime('%Y-%m-%d') if check.delivered_at else None,
            'paid_at': check.paid_at.strftime('%Y-%m-%d') if check.paid_at else None,
        } for check in checks]

        return JsonResponse({
            'payment_details': payment_details,
            'checks': check_details
        })

class InvoiceAccountingSummaryView(View):
    def get(self, request, invoice_id):
        invoice = get_object_or_404(Invoice, id=invoice_id)
        
        # Get original entries
        original_entries = invoice.get_accounting_entries()
        
        # Get credit note entries
        credit_note_entries = []
        credit_notes_total = 0
        for credit_note in invoice.credit_notes.all():
            entries = credit_note.get_accounting_entries()
            credit_note_entries.extend(entries)
            credit_notes_total += credit_note.total_amount
            
        return JsonResponse({
            'original_entries': original_entries,
            'credit_note_entries': credit_note_entries,
            'totals': {
                'original': float(invoice.total_amount),
                'credit_notes': float(-credit_notes_total),
                'net': float(invoice.net_amount)
            }
        })